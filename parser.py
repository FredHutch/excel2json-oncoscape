from openpyxl import load_workbook
import subprocess
import os
# wb = load_workbook(filename='GBM-LGG.xlsx')
wb = load_workbook(filename='GBM-LGG.xlsx', read_only=True)
# wb = load_workbook(filename='uploading_demo.xlsx', read_only=True)
sheetNames = wb.sheetnames;

for sheetName in sheetNames:
    print('***')
    print(sheetName)
    command_string = "xlsx2csv GBM-LGG.xlsx -n " + sheetName + " > " + sheetName + ".csv"
    os.system(command_string)
    # print subprocess.check_call(['ls', '-l'])    
